#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
大乐透预测（多算法集成）

玩法：前区 01-35 选5个；后区 01-12 选2个；顺序无关。
算法侧重：频率/遗漏/形态/共现/贝叶斯/EMA/间隔/蒙特卡洛（适合无序组合）。
"""

from __future__ import annotations

import itertools
import json
import math
import random
import shutil
import sys
from collections import Counter, defaultdict
from pathlib import Path
from typing import Dict, List, Optional, Sequence, Tuple

FRONT_MAX = 35
BACK_MAX = 12
FRONT_PICK = 5
BACK_PICK = 2


def app_root() -> Path:
    if getattr(sys, "frozen", False):
        return Path(sys.executable).resolve().parent
    # lottery_suite/ is parent of pl3/ or dlt/
    return Path(__file__).resolve().parent.parent


def bundled_data_dir() -> Path | None:
    if getattr(sys, "frozen", False) and hasattr(sys, "_MEIPASS"):
        return Path(sys._MEIPASS) / "data" / "dlt"
    return None


ROOT = app_root()
DATA_DIR = ROOT / "data" / "dlt"
DATA_PATH = DATA_DIR / "history.json"
MODEL_PATH = DATA_DIR / "model.json"
LAST_PRED_PATH = DATA_DIR / "last_prediction.json"
COMPARE_LOG_PATH = DATA_DIR / "compare_log.json"
WEIGHTS_PATH = DATA_DIR / "algo_weights.json"

# 号码融合算法 / 组合打分算法
NUMBER_KEYS = ["freq", "gap", "ema", "bayes", "skip", "cooc"]
COMBO_KEYS = ["number", "pattern", "cooccur", "markov", "gap_avg"]
ALGO_NAMES = {
    "freq": "加权频率",
    "gap": "遗漏冷热",
    "ema": "EMA趋势",
    "bayes": "贝叶斯融合",
    "skip": "间隔规律",
    "cooc": "共现关联",
    "number": "号码融合分",
    "pattern": "形态匹配",
    "cooccur": "共现加分",
    "markov": "集合马尔可夫",
    "gap_avg": "冷热组合分",
}
DEFAULT_NUMBER_WEIGHTS = {
    "freq": 0.18,
    "gap": 0.22,
    "ema": 0.14,
    "bayes": 0.16,
    "skip": 0.14,
    "cooc": 0.16,
}
DEFAULT_COMBO_WEIGHTS = {
    "number": 0.42,
    "pattern": 0.22,
    "cooccur": 0.18,
    "markov": 0.10,
    "gap_avg": 0.08,
}

Draw = Dict[str, object]
Front = Tuple[int, ...]
Back = Tuple[int, ...]


def ensure_data_files() -> None:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    if DATA_PATH.exists():
        return
    seed = bundled_data_dir()
    if seed and (seed / "history.json").exists():
        shutil.copy2(seed / "history.json", DATA_PATH)


def _norm_nums(nums: Sequence[int], lo: int, hi: int, need: int) -> List[int]:
    out = sorted(int(x) for x in nums)
    if len(out) != need or len(set(out)) != need:
        raise ValueError(f"需要 {need} 个不重复号码，收到 {nums}")
    if any(x < lo or x > hi for x in out):
        raise ValueError(f"号码需在 {lo:02d}-{hi:02d}：{out}")
    return out


def load_history(path: Path = DATA_PATH) -> List[Draw]:
    ensure_data_files()
    if not path.exists():
        return []
    with path.open("r", encoding="utf-8") as f:
        raw = json.load(f)
    draws: List[Draw] = []
    for item in raw:
        front = _norm_nums(item["front"], 1, FRONT_MAX, FRONT_PICK)
        back = _norm_nums(item["back"], 1, BACK_MAX, BACK_PICK)
        draws.append(
            {
                "period": int(item["period"]),
                "date": str(item.get("date", "")),
                "front": front,
                "back": back,
            }
        )
    draws.sort(key=lambda x: int(x["period"]))
    return draws


def save_history(draws: Sequence[Draw], path: Path = DATA_PATH) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    ordered = sorted(draws, key=lambda x: int(x["period"]))
    payload = []
    for d in ordered:
        payload.append(
            {
                "period": int(d["period"]),
                "date": d.get("date", ""),
                "front": list(d["front"]),
                "back": list(d["back"]),
            }
        )
    with path.open("w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)


def add_draw(
    period: int,
    front: Sequence[int],
    back: Sequence[int],
    date: str = "",
) -> List[Draw]:
    draws = load_history()
    for d in draws:
        if int(d["period"]) == period:
            raise ValueError(f"期号 {period} 已存在")
    draws.append(
        {
            "period": period,
            "date": date,
            "front": _norm_nums(front, 1, FRONT_MAX, FRONT_PICK),
            "back": _norm_nums(back, 1, BACK_MAX, BACK_PICK),
        }
    )
    save_history(draws)
    return draws


def delete_draw(period: int) -> List[Draw]:
    draws = load_history()
    kept = [d for d in draws if int(d["period"]) != period]
    if len(kept) == len(draws):
        raise ValueError(f"未找到期号 {period}")
    save_history(kept)
    return kept


def delete_latest() -> Tuple[Draw, List[Draw]]:
    draws = load_history()
    if not draws:
        raise ValueError("没有可删除的数据")
    removed = draws[-1]
    kept = draws[:-1]
    save_history(kept)
    return removed, kept


def clear_all_history() -> None:
    save_history([])


def fmt_nums(nums: Sequence[int]) -> str:
    return " ".join(f"{int(x):02d}" for x in nums)


def normalize(scores: Dict[int, float]) -> Dict[int, float]:
    s = sum(max(0.0, v) for v in scores.values())
    if s <= 0:
        n = len(scores) or 1
        return {k: 1.0 / n for k in scores}
    return {k: max(0.0, v) / s for k, v in scores.items()}


def softmax_dict(scores: Dict[object, float], temperature: float = 1.0) -> Dict[object, float]:
    if not scores:
        return {}
    t = max(temperature, 1e-6)
    mx = max(scores.values())
    exps = {k: math.exp((v - mx) / t) for k, v in scores.items()}
    s = sum(exps.values()) or 1.0
    return {k: v / s for k, v in exps.items()}


# ---------------------------------------------------------------------------
# 号码级模型（前区 1-35 / 后区 1-12）
# ---------------------------------------------------------------------------

class WeightedFrequency:
    def __init__(self, half_life: float = 25.0):
        self.half_life = half_life
        self.front: Dict[int, float] = {}
        self.back: Dict[int, float] = {}

    def fit(self, draws: Sequence[Draw]) -> None:
        fc = Counter()
        bc = Counter()
        n = len(draws)
        for i, d in enumerate(draws):
            w = 0.5 ** ((n - 1 - i) / self.half_life)
            for x in d["front"]:
                fc[int(x)] += w
            for x in d["back"]:
                bc[int(x)] += w
        self.front = normalize({k: fc.get(k, 0) + 0.3 for k in range(1, FRONT_MAX + 1)})
        self.back = normalize({k: bc.get(k, 0) + 0.3 for k in range(1, BACK_MAX + 1)})


class GapHotCold:
    """遗漏 + 近期热度，大乐透最常用统计法。"""

    def fit(self, draws: Sequence[Draw]) -> None:
        n = len(draws)
        window = min(30, n)
        self.fgap = {k: n for k in range(1, FRONT_MAX + 1)}
        self.bgap = {k: n for k in range(1, BACK_MAX + 1)}
        self.fhot = Counter()
        self.bhot = Counter()
        for i, d in enumerate(draws):
            for x in d["front"]:
                self.fgap[int(x)] = n - 1 - i
                if i >= n - window:
                    self.fhot[int(x)] += 1
            for x in d["back"]:
                self.bgap[int(x)] = n - 1 - i
                if i >= n - window:
                    self.bhot[int(x)] += 1
        self.front = normalize(
            {
                k: 1.0 + 0.12 * min(self.fgap[k], 18) + 0.4 * self.fhot.get(k, 0)
                for k in range(1, FRONT_MAX + 1)
            }
        )
        self.back = normalize(
            {
                k: 1.0 + 0.15 * min(self.bgap[k], 12) + 0.45 * self.bhot.get(k, 0)
                for k in range(1, BACK_MAX + 1)
            }
        )


class EMATrend:
    def __init__(self, alpha: float = 0.2):
        self.alpha = alpha
        self.front: Dict[int, float] = {}
        self.back: Dict[int, float] = {}

    def fit(self, draws: Sequence[Draw]) -> None:
        a = self.alpha
        fe = {k: 1.0 / FRONT_MAX for k in range(1, FRONT_MAX + 1)}
        be = {k: 1.0 / BACK_MAX for k in range(1, BACK_MAX + 1)}
        for d in draws:
            fs, bs = set(map(int, d["front"])), set(map(int, d["back"]))
            for k in fe:
                fe[k] = a * (1.0 if k in fs else 0.0) + (1 - a) * fe[k]
            for k in be:
                be[k] = a * (1.0 if k in bs else 0.0) + (1 - a) * be[k]
        self.front = normalize(fe)
        self.back = normalize(be)


class BayesianFusion:
    def __init__(self, prior: float = 1.0):
        self.prior = prior
        self.front: Dict[int, float] = {}
        self.back: Dict[int, float] = {}

    def fit(self, draws: Sequence[Draw]) -> None:
        fc, bc = Counter(), Counter()
        for i, d in enumerate(draws):
            w = 1.0 + 0.015 * i
            for x in d["front"]:
                fc[int(x)] += w
            for x in d["back"]:
                bc[int(x)] += w
        self.front = normalize(
            {k: self.prior + fc.get(k, 0) for k in range(1, FRONT_MAX + 1)}
        )
        self.back = normalize(
            {k: self.prior + bc.get(k, 0) for k in range(1, BACK_MAX + 1)}
        )


class SkipInterval:
    def fit(self, draws: Sequence[Draw]) -> None:
        n = len(draws)
        last_f = {k: None for k in range(1, FRONT_MAX + 1)}
        last_b = {k: None for k in range(1, BACK_MAX + 1)}
        pref_f = [Counter() for _ in range(FRONT_MAX + 1)]
        pref_b = [Counter() for _ in range(BACK_MAX + 1)]
        for i, d in enumerate(draws):
            for x in map(int, d["front"]):
                if last_f[x] is not None:
                    pref_f[x][i - last_f[x]] += 1
                last_f[x] = i
            for x in map(int, d["back"]):
                if last_b[x] is not None:
                    pref_b[x][i - last_b[x]] += 1
                last_b[x] = i
        cur_f = {k: (n - last_f[k]) if last_f[k] is not None else n for k in last_f}
        cur_b = {k: (n - last_b[k]) if last_b[k] is not None else n for k in last_b}

        def score(pref: Counter, gap: int) -> float:
            total = sum(pref.values()) or 1
            return 0.15 + pref.get(gap, 0) / total + 0.04 * pref.get(gap - 1, 0) / total + 0.04 * pref.get(gap + 1, 0) / total

        self.front = normalize({k: score(pref_f[k], cur_f[k]) for k in range(1, FRONT_MAX + 1)})
        self.back = normalize({k: score(pref_b[k], cur_b[k]) for k in range(1, BACK_MAX + 1)})


class CoOccurrence:
    """共现：历史上常一起出现的号码互相抬分（适合无序组合）。"""

    def fit(self, draws: Sequence[Draw]) -> None:
        self.fpair = Counter()
        self.bpair = Counter()
        self.last_front: Front = ()
        self.last_back: Back = ()
        for d in draws:
            f = tuple(sorted(map(int, d["front"])))
            b = tuple(sorted(map(int, d["back"])))
            for a, c in itertools.combinations(f, 2):
                self.fpair[(a, c)] += 1
            if len(b) == 2:
                self.bpair[tuple(sorted(b))] += 1
            self.last_front = f
            self.last_back = b

        # 以最近一期为锚，给与其共现高的号码加分
        fscore = {k: 0.5 for k in range(1, FRONT_MAX + 1)}
        for x in self.last_front:
            fscore[x] += 0.2
            for y in range(1, FRONT_MAX + 1):
                if y == x:
                    continue
                key = tuple(sorted((x, y)))
                fscore[y] += 0.08 * self.fpair.get(key, 0)
        bscore = {k: 0.5 for k in range(1, BACK_MAX + 1)}
        for x in self.last_back:
            bscore[x] += 0.2
            for y in range(1, BACK_MAX + 1):
                if y == x:
                    continue
                key = tuple(sorted((x, y)))
                bscore[y] += 0.15 * self.bpair.get(key, 0)
        self.front = normalize(fscore)
        self.back = normalize(bscore)

    def pair_bonus(self, front: Sequence[int], back: Sequence[int]) -> float:
        f = sorted(map(int, front))
        b = sorted(map(int, back))
        fp = sum(self.fpair.get(tuple(sorted(p)), 0) for p in itertools.combinations(f, 2))
        bp = self.bpair.get(tuple(b), 0)
        # 归一到大致 0~1
        return math.log1p(fp) / 8.0 + math.log1p(bp) / 4.0


class PatternModel:
    """形态：和值、奇偶、区间分布（前区）。"""

    def fit(self, draws: Sequence[Draw]) -> None:
        self.sum_f = Counter()
        self.oe_f = Counter()  # odd count
        self.zone_f = Counter()  # (low1-12, mid13-24, high25-35) as tuple
        self.sum_b = Counter()
        self.oe_b = Counter()
        n = len(draws)
        for i, d in enumerate(draws):
            w = 0.5 ** ((n - 1 - i) / 30.0)
            f = list(map(int, d["front"]))
            b = list(map(int, d["back"]))
            self.sum_f[sum(f)] += w
            self.oe_f[sum(x % 2 for x in f)] += w
            z = (
                sum(1 for x in f if x <= 12),
                sum(1 for x in f if 13 <= x <= 24),
                sum(1 for x in f if x >= 25),
            )
            self.zone_f[z] += w
            self.sum_b[sum(b)] += w
            self.oe_b[sum(x % 2 for x in b)] += w
        self.sum_fp = normalize({k: float(v) for k, v in self.sum_f.items()})
        self.oe_fp = normalize({k: float(v) for k, v in self.oe_f.items()})
        self.zone_fp = normalize({k: float(v) for k, v in self.zone_f.items()})
        self.sum_bp = normalize({k: float(v) for k, v in self.sum_b.items()})
        self.oe_bp = normalize({k: float(v) for k, v in self.oe_b.items()})

    def score(self, front: Sequence[int], back: Sequence[int]) -> float:
        f = list(map(int, front))
        b = list(map(int, back))
        z = (
            sum(1 for x in f if x <= 12),
            sum(1 for x in f if 13 <= x <= 24),
            sum(1 for x in f if x >= 25),
        )
        return (
            self.sum_fp.get(sum(f), 1e-4)
            * self.oe_fp.get(sum(x % 2 for x in f), 1e-4)
            * self.zone_fp.get(z, 1e-4)
            * self.sum_bp.get(sum(b), 1e-4)
            * self.oe_bp.get(sum(x % 2 for x in b), 1e-4)
        )


class MarkovSet:
    """把每期前/后区排序后做成状态，做整注转移（稀疏，作弱信号）。"""

    def fit(self, draws: Sequence[Draw]) -> None:
        self.f_trans: Dict[str, Counter] = defaultdict(Counter)
        self.b_trans: Dict[str, Counter] = defaultdict(Counter)
        self.f_global = Counter()
        self.b_global = Counter()
        self.last_f = ""
        self.last_b = ""
        keys_f = ["-".join(f"{x:02d}" for x in d["front"]) for d in draws]
        keys_b = ["-".join(f"{x:02d}" for x in d["back"]) for d in draws]
        for i in range(len(keys_f) - 1):
            self.f_trans[keys_f[i]][keys_f[i + 1]] += 1
            self.b_trans[keys_b[i]][keys_b[i + 1]] += 1
            self.f_global[keys_f[i + 1]] += 1
            self.b_global[keys_b[i + 1]] += 1
        if keys_f:
            self.last_f = keys_f[-1]
            self.last_b = keys_b[-1]

    def score(self, front: Sequence[int], back: Sequence[int]) -> float:
        fk = "-".join(f"{int(x):02d}" for x in sorted(front))
        bk = "-".join(f"{int(x):02d}" for x in sorted(back))
        fs = self.f_trans.get(self.last_f, Counter()).get(fk, 0) + 0.05 * self.f_global.get(fk, 0)
        bs = self.b_trans.get(self.last_b, Counter()).get(bk, 0) + 0.05 * self.b_global.get(bk, 0)
        return (fs + 0.01) * (bs + 0.01)


# ---------------------------------------------------------------------------
# 集成
# ---------------------------------------------------------------------------

def load_algo_weights(path: Path = WEIGHTS_PATH) -> dict:
    """读取可自适应修正的算法权重。"""
    data = {
        "number": dict(DEFAULT_NUMBER_WEIGHTS),
        "combo": dict(DEFAULT_COMBO_WEIGHTS),
        "updates": 0,
        "history": [],
    }
    if path.exists():
        try:
            with path.open("r", encoding="utf-8") as f:
                raw = json.load(f)
            for k in NUMBER_KEYS:
                if k in raw.get("number", {}):
                    data["number"][k] = float(raw["number"][k])
            for k in COMBO_KEYS:
                if k in raw.get("combo", {}):
                    data["combo"][k] = float(raw["combo"][k])
            data["updates"] = int(raw.get("updates", 0))
            data["history"] = list(raw.get("history", []))[-50:]
        except Exception:
            pass
    data["number"] = _clamp_normalize(data["number"], min_w=0.05)
    data["combo"] = _clamp_normalize(data["combo"], min_w=0.04)
    return data


def save_algo_weights(data: dict, path: Path = WEIGHTS_PATH) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def _clamp_normalize(weights: Dict[str, float], min_w: float = 0.05) -> Dict[str, float]:
    keys = list(weights.keys())
    clamped = {k: max(min_w, float(weights[k])) for k in keys}
    s = sum(clamped.values()) or 1.0
    return {k: v / s for k, v in clamped.items()}


def reset_algo_weights(path: Path = WEIGHTS_PATH) -> dict:
    data = {
        "number": dict(DEFAULT_NUMBER_WEIGHTS),
        "combo": dict(DEFAULT_COMBO_WEIGHTS),
        "updates": 0,
        "history": [],
    }
    save_algo_weights(data, path)
    return data


class EnsemblePredictor:
    """
    适合大乐透的算法：
    1 加权频率  2 遗漏冷热  3 EMA  4 贝叶斯  5 间隔  6 共现
    7 形态匹配  8 集合马尔可夫  9 蒙特卡洛采样
    权重会在每次「录入对比」后按准确率高低自动修正。
    """

    def __init__(self, weights: Optional[dict] = None):
        self.freq = WeightedFrequency()
        self.gap = GapHotCold()
        self.ema = EMATrend()
        self.bayes = BayesianFusion()
        self.skip = SkipInterval()
        self.cooc = CoOccurrence()
        self.pattern = PatternModel()
        self.markov = MarkovSet()
        self.draws: List[Draw] = []
        self.front_blend: Dict[int, float] = {}
        self.back_blend: Dict[int, float] = {}
        self.weights = weights if weights is not None else load_algo_weights()

    def fit(self, draws: Sequence[Draw]) -> None:
        self.draws = list(draws)
        for m in (self.freq, self.gap, self.ema, self.bayes, self.skip, self.cooc, self.pattern, self.markov):
            m.fit(draws)

        src_map_f = {
            "freq": self.freq.front,
            "gap": self.gap.front,
            "ema": self.ema.front,
            "bayes": self.bayes.front,
            "skip": self.skip.front,
            "cooc": self.cooc.front,
        }
        src_map_b = {
            "freq": self.freq.back,
            "gap": self.gap.back,
            "ema": self.ema.back,
            "bayes": self.bayes.back,
            "skip": self.skip.back,
            "cooc": self.cooc.back,
        }
        nw = self.weights["number"]
        fb = {k: 0.0 for k in range(1, FRONT_MAX + 1)}
        bb = {k: 0.0 for k in range(1, BACK_MAX + 1)}
        for key in NUMBER_KEYS:
            w = nw[key]
            for k in fb:
                fb[k] += w * src_map_f[key][k]
            for k in bb:
                bb[k] += w * src_map_b[key][k]
        self.front_blend = normalize(fb)
        self.back_blend = normalize(bb)

    def _candidates(self) -> List[Tuple[Front, Back]]:
        top_f = sorted(self.front_blend, key=self.front_blend.get, reverse=True)
        top_b = sorted(self.back_blend, key=self.back_blend.get, reverse=True)
        pool: set[Tuple[Front, Back]] = set()

        # 前区 Top12 取5 + 后区 Top6 取2
        for f in itertools.combinations(top_f[:12], FRONT_PICK):
            for b in itertools.combinations(top_b[:6], BACK_PICK):
                pool.add((tuple(sorted(f)), tuple(sorted(b))))

        # Top15 取5 × Top8 取2（更大覆盖，限制数量）
        for f in itertools.combinations(top_f[:15], FRONT_PICK):
            for b in itertools.combinations(top_b[:5], BACK_PICK):
                pool.add((tuple(sorted(f)), tuple(sorted(b))))
                if len(pool) > 2500:
                    break
            if len(pool) > 2500:
                break

        # 历史整注
        for d in self.draws:
            pool.add((tuple(d["front"]), tuple(d["back"])))

        # 蒙特卡洛采样
        rng = random.Random(42)
        for _ in range(600):
            f = self._sample_without_replacement(self.front_blend, FRONT_PICK, rng)
            b = self._sample_without_replacement(self.back_blend, BACK_PICK, rng)
            pool.add((tuple(sorted(f)), tuple(sorted(b))))

        return list(pool)

    @staticmethod
    def _sample_without_replacement(probs: Dict[int, float], k: int, rng: random.Random) -> List[int]:
        items = list(probs.keys())
        weights = [max(1e-12, probs[i]) for i in items]
        chosen = []
        pool_i = items[:]
        pool_w = weights[:]
        for _ in range(k):
            s = sum(pool_w) or 1.0
            r = rng.random() * s
            acc = 0.0
            idx = 0
            for j, w in enumerate(pool_w):
                acc += w
                if r <= acc:
                    idx = j
                    break
            chosen.append(pool_i[idx])
            del pool_i[idx]
            del pool_w[idx]
        return chosen

    def _combo_number_score(self, front: Front, back: Back) -> float:
        fs = sum(math.log(self.front_blend[x] + 1e-12) for x in front)
        bs = sum(math.log(self.back_blend[x] + 1e-12) for x in back)
        return fs + 1.15 * bs

    def predict(self, top_n: int = 10) -> List[Tuple[Front, Back, float, Dict[str, float]]]:
        if len(self.draws) < 10:
            raise RuntimeError("历史数据太少，至少需要 10 期")

        cands = self._candidates()
        scored = []
        for front, back in cands:
            detail = {
                "number": self._combo_number_score(front, back),
                "pattern": self.pattern.score(front, back),
                "cooccur": self.cooc.pair_bonus(front, back),
                "markov": self.markov.score(front, back),
                "gap_avg": (
                    sum(self.gap.front[x] for x in front) / FRONT_PICK
                    + sum(self.gap.back[x] for x in back) / BACK_PICK
                ) / 2,
            }
            cw = self.weights["combo"]
            score = (
                cw["number"] * detail["number"]
                + cw["pattern"] * math.log(detail["pattern"] + 1e-12)
                + cw["cooccur"] * detail["cooccur"]
                + cw["markov"] * math.log(detail["markov"] + 1e-12)
                + cw["gap_avg"] * math.log(detail["gap_avg"] + 1e-12)
            )
            scored.append((front, back, score, detail))

        score_map = {i: s for i, (_, _, s, _) in enumerate(scored)}
        prob_map = softmax_dict(score_map, temperature=1.1)
        ranked = sorted(
            ((scored[i][0], scored[i][1], prob_map[i], scored[i][3]) for i in range(len(scored))),
            key=lambda x: x[2],
            reverse=True,
        )
        # 多样性：前区至少差 2 个号
        picked: List[Tuple[Front, Back, float, Dict[str, float]]] = []
        for item in ranked:
            fset = set(item[0])
            if any(len(fset & set(p[0])) >= 4 for p in picked):
                continue
            picked.append(item)
            if len(picked) >= top_n:
                break
        while len(picked) < min(top_n, len(ranked)):
            for item in ranked:
                if item not in picked:
                    picked.append(item)
                if len(picked) >= top_n:
                    break
        return picked[:top_n]

    def hot_cold_boards(self, top_k: int = 12) -> Dict[str, List[Tuple[int, float]]]:
        front_sorted = sorted(self.front_blend.items(), key=lambda x: x[1], reverse=True)
        back_sorted = sorted(self.back_blend.items(), key=lambda x: x[1], reverse=True)
        return {
            "front_hot": front_sorted[:top_k],
            "front_cold": list(reversed(front_sorted[-top_k:])),
            "back_hot": back_sorted[:top_k],
            "back_cold": list(reversed(back_sorted[-min(top_k, BACK_MAX):])),
        }

    def save(self, path: Path = MODEL_PATH) -> None:
        path.parent.mkdir(parents=True, exist_ok=True)
        meta = {
            "n_draws": len(self.draws),
            "last_period": self.draws[-1]["period"] if self.draws else None,
            "last_front": list(self.draws[-1]["front"]) if self.draws else None,
            "last_back": list(self.draws[-1]["back"]) if self.draws else None,
            "weights": self.weights,
            "algorithms": [
                "加权频率", "遗漏冷热", "EMA趋势", "贝叶斯融合",
                "间隔规律", "共现关联", "形态匹配", "集合马尔可夫", "蒙特卡洛采样",
            ],
        }
        with path.open("w", encoding="utf-8") as f:
            json.dump(meta, f, ensure_ascii=False, indent=2)

    def number_sources(self) -> Dict[str, Tuple[Dict[int, float], Dict[int, float]]]:
        return {
            "freq": (self.freq.front, self.freq.back),
            "gap": (self.gap.front, self.gap.back),
            "ema": (self.ema.front, self.ema.back),
            "bayes": (self.bayes.front, self.bayes.back),
            "skip": (self.skip.front, self.skip.back),
            "cooc": (self.cooc.front, self.cooc.back),
        }


# ---------------------------------------------------------------------------
# 预测存档 & 开奖对比（Top10 命中）
# ---------------------------------------------------------------------------

def save_last_prediction(
    target_period: int,
    preds: Sequence[Tuple[Front, Back, float, Dict[str, float]]],
    based_on_period: Optional[int] = None,
    path: Path = LAST_PRED_PATH,
) -> None:
    """保存本次预测 Top 结果，供下一期开奖后对比。"""
    ensure_data_files()
    top = []
    for i, (front, back, prob, _) in enumerate(preds[:10], 1):
        top.append(
            {
                "rank": i,
                "front": list(front),
                "back": list(back),
                "prob": float(prob),
            }
        )
    payload = {
        "target_period": int(target_period),
        "based_on_period": based_on_period,
        "top": top,
    }
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)


def load_last_prediction(path: Path = LAST_PRED_PATH) -> Optional[dict]:
    if not path.exists():
        return None
    with path.open("r", encoding="utf-8") as f:
        return json.load(f)


def _score_number_algo(
    front_p: Dict[int, float],
    back_p: Dict[int, float],
    actual_front: Sequence[int],
    actual_back: Sequence[int],
) -> dict:
    """用该算法 Top5前区 + Top2后区 对照开奖，计算命中准确率。"""
    af = set(map(int, actual_front))
    ab = set(map(int, actual_back))
    top_f = sorted(front_p, key=front_p.get, reverse=True)[:FRONT_PICK]
    top_b = sorted(back_p, key=back_p.get, reverse=True)[:BACK_PICK]
    fh = len(af & set(top_f))
    bh = len(ab & set(top_b))
    like = sum(front_p.get(x, 0) for x in af) + sum(back_p.get(x, 0) for x in ab)
    return {
        "front_hit": fh,
        "back_hit": bh,
        "total_hit": fh + bh,
        "accuracy": (fh + bh) / 7.0,
        "likelihood": like,
        "pick_front": top_f,
        "pick_back": top_b,
    }


def _score_combo_algo_top10(
    cand_scores: List[Tuple[Front, Back, float]],
    actual_front: Sequence[int],
    actual_back: Sequence[int],
) -> dict:
    """单算法按分数取 Top10，看最佳一注命中。"""
    af = set(map(int, actual_front))
    ab = set(map(int, actual_back))
    ranked = sorted(cand_scores, key=lambda x: x[2], reverse=True)[:10]
    best_total = -1
    best = None
    totals = []
    for front, back, _ in ranked:
        fh = len(af & set(front))
        bh = len(ab & set(back))
        tot = fh + bh
        totals.append(tot)
        if tot > best_total:
            best_total = tot
            best = {"front": list(front), "back": list(back), "front_hit": fh, "back_hit": bh}
    avg = sum(totals) / len(totals) if totals else 0.0
    return {
        "front_hit": best["front_hit"] if best else 0,
        "back_hit": best["back_hit"] if best else 0,
        "total_hit": best_total if best_total >= 0 else 0,
        "accuracy": (best_total / 7.0) if best_total >= 0 else 0.0,
        "avg_total_hit": avg,
        "likelihood": avg,
        "pick_front": best["front"] if best else [],
        "pick_back": best["back"] if best else [],
    }


def _target_weights_from_ranking(scores: Dict[str, dict], keys: Sequence[str]) -> Dict[str, float]:
    """按准确率从高到低排名，生成目标权重（准确率 + 名次双信号）。"""
    ordered = sorted(
        keys,
        key=lambda k: (scores[k]["accuracy"], scores[k].get("likelihood", 0.0)),
        reverse=True,
    )
    n = len(ordered)
    rank_raw = {k: float(n - i) for i, k in enumerate(ordered)}
    acc_raw = {k: scores[k]["accuracy"] + 0.08 for k in keys}
    rank_w = normalize(rank_raw)
    acc_w = normalize(acc_raw)
    mixed = {k: 0.55 * acc_w[k] + 0.45 * rank_w[k] for k in keys}
    return normalize(mixed)


def correct_algorithms_from_draw(
    actual_front: Sequence[int],
    actual_back: Sequence[int],
    actual_period: Optional[int] = None,
    ema_alpha: float = 0.40,
    force: bool = False,
) -> dict:
    """
    用本期开奖回测各算法，按准确率从高到低修正权重并持久化。
    训练数据 = 本期之前的历史。同一期默认只修正一次。
    """
    old = load_algo_weights()
    already = (
        not force
        and actual_period is not None
        and bool(old.get("history"))
        and old["history"][-1].get("period") == actual_period
    )

    history = load_history()
    if actual_period is not None:
        train = [d for d in history if int(d["period"]) < int(actual_period)]
    else:
        train = history[:-1] if history else []
    if len(train) < 10:
        raise ValueError("用于算法修正的历史不足 10 期。")

    # 用默认权重拟合，评估各子算法原始表现（避免被旧权重污染）
    model = EnsemblePredictor(
        weights={
            "number": dict(DEFAULT_NUMBER_WEIGHTS),
            "combo": dict(DEFAULT_COMBO_WEIGHTS),
            "updates": 0,
            "history": [],
        }
    )
    model.fit(train)

    number_scores = {}
    for key, (fp, bp) in model.number_sources().items():
        number_scores[key] = _score_number_algo(fp, bp, actual_front, actual_back)

    cands = model._candidates()
    combo_raw: Dict[str, List[Tuple[Front, Back, float]]] = {k: [] for k in COMBO_KEYS}
    for front, back in cands:
        detail = {
            "number": model._combo_number_score(front, back),
            "pattern": model.pattern.score(front, back),
            "cooccur": model.cooc.pair_bonus(front, back),
            "markov": model.markov.score(front, back),
            "gap_avg": (
                sum(model.gap.front[x] for x in front) / FRONT_PICK
                + sum(model.gap.back[x] for x in back) / BACK_PICK
            ) / 2,
        }
        combo_raw["number"].append((front, back, detail["number"]))
        combo_raw["pattern"].append((front, back, math.log(detail["pattern"] + 1e-12)))
        combo_raw["cooccur"].append((front, back, detail["cooccur"]))
        combo_raw["markov"].append((front, back, math.log(detail["markov"] + 1e-12)))
        combo_raw["gap_avg"].append((front, back, math.log(detail["gap_avg"] + 1e-12)))

    combo_scores = {
        k: _score_combo_algo_top10(combo_raw[k], actual_front, actual_back) for k in COMBO_KEYS
    }

    number_rank = sorted(
        NUMBER_KEYS,
        key=lambda k: (number_scores[k]["accuracy"], number_scores[k]["likelihood"]),
        reverse=True,
    )
    combo_rank = sorted(
        COMBO_KEYS,
        key=lambda k: (combo_scores[k]["accuracy"], combo_scores[k]["likelihood"]),
        reverse=True,
    )

    if already:
        new_number = dict(old["number"])
        new_combo = dict(old["combo"])
        updates = int(old.get("updates", 0))
    else:
        target_n = _target_weights_from_ranking(number_scores, NUMBER_KEYS)
        target_c = _target_weights_from_ranking(combo_scores, COMBO_KEYS)
        new_number = {
            k: (1 - ema_alpha) * old["number"][k] + ema_alpha * target_n[k] for k in NUMBER_KEYS
        }
        new_combo = {
            k: (1 - ema_alpha) * old["combo"][k] + ema_alpha * target_c[k] for k in COMBO_KEYS
        }
        new_number = _clamp_normalize(new_number, min_w=0.05)
        new_combo = _clamp_normalize(new_combo, min_w=0.04)
        updates = int(old.get("updates", 0)) + 1
        payload = {
            "number": new_number,
            "combo": new_combo,
            "updates": updates,
            "history": list(old.get("history", [])),
        }
        payload["history"].append(
            {
                "period": actual_period,
                "number_rank": [ALGO_NAMES[k] for k in number_rank],
                "combo_rank": [ALGO_NAMES[k] for k in combo_rank],
                "number_weights": new_number,
                "combo_weights": new_combo,
            }
        )
        payload["history"] = payload["history"][-50:]
        save_algo_weights(payload)

    ranking_rows = []
    for i, k in enumerate(number_rank, 1):
        ranking_rows.append(
            {
                "group": "号码算法",
                "rank": i,
                "key": k,
                "name": ALGO_NAMES[k],
                "accuracy": number_scores[k]["accuracy"],
                "total_hit": number_scores[k]["total_hit"],
                "front_hit": number_scores[k]["front_hit"],
                "back_hit": number_scores[k]["back_hit"],
                "old_weight": old["number"][k],
                "new_weight": new_number[k],
                "delta": new_number[k] - old["number"][k],
            }
        )
    for i, k in enumerate(combo_rank, 1):
        ranking_rows.append(
            {
                "group": "组合算法",
                "rank": i,
                "key": k,
                "name": ALGO_NAMES[k],
                "accuracy": combo_scores[k]["accuracy"],
                "total_hit": combo_scores[k]["total_hit"],
                "front_hit": combo_scores[k]["front_hit"],
                "back_hit": combo_scores[k]["back_hit"],
                "old_weight": old["combo"][k],
                "new_weight": new_combo[k],
                "delta": new_combo[k] - old["combo"][k],
            }
        )

    return {
        "number_rank": number_rank,
        "combo_rank": combo_rank,
        "number_scores": number_scores,
        "combo_scores": combo_scores,
        "ranking_rows": ranking_rows,
        "old_weights": {"number": old["number"], "combo": old["combo"]},
        "new_weights": {"number": new_number, "combo": new_combo},
        "updates": updates,
        "skipped_duplicate": already,
    }


def compare_prediction_to_draw(
    actual_front: Sequence[int],
    actual_back: Sequence[int],
    prediction: Optional[dict] = None,
    actual_period: Optional[int] = None,
    correct_weights: bool = True,
) -> dict:
    """
    用开奖号码对比上次预测 Top10。
    每注统计：前区命中(0-5)、后区命中(0-2)、合计(0-7)、准确率。
    默认同时按各算法准确率高低修正集成权重。
    """
    pred = prediction if prediction is not None else load_last_prediction()
    if not pred or not pred.get("top"):
        raise ValueError("没有可对比的上次预测，请先点击「开始预测」生成 Top10。")

    af = set(map(int, actual_front))
    ab = set(map(int, actual_back))
    rows = []
    for item in pred["top"][:10]:
        pf = set(map(int, item["front"]))
        pb = set(map(int, item["back"]))
        front_hit = len(af & pf)
        back_hit = len(ab & pb)
        total_hit = front_hit + back_hit
        rows.append(
            {
                "rank": int(item["rank"]),
                "front": sorted(pf),
                "back": sorted(pb),
                "front_hit": front_hit,
                "back_hit": back_hit,
                "total_hit": total_hit,
                "accuracy": total_hit / 7.0,
                "front_hit_nums": sorted(af & pf),
                "back_hit_nums": sorted(ab & pb),
            }
        )

    best = max(rows, key=lambda r: (r["total_hit"], r["back_hit"], -r["rank"]))
    avg_front = sum(r["front_hit"] for r in rows) / len(rows)
    avg_back = sum(r["back_hit"] for r in rows) / len(rows)
    avg_total = sum(r["total_hit"] for r in rows) / len(rows)
    avg_acc = sum(r["accuracy"] for r in rows) / len(rows)

    pool_f = set()
    pool_b = set()
    for r in rows:
        pool_f.update(r["front"])
        pool_b.update(r["back"])
    cover_front = len(af & pool_f)
    cover_back = len(ab & pool_b)

    period_match = True
    target_period = pred.get("target_period")
    if actual_period is not None and target_period is not None:
        period_match = int(actual_period) == int(target_period)

    correction = None
    if correct_weights:
        try:
            correction = correct_algorithms_from_draw(
                actual_front, actual_back, actual_period=actual_period
            )
        except ValueError:
            correction = None

    result = {
        "target_period": target_period,
        "actual_period": actual_period,
        "period_match": period_match,
        "actual_front": sorted(af),
        "actual_back": sorted(ab),
        "rows": rows,
        "best": best,
        "summary": {
            "top_n": len(rows),
            "avg_front_hit": avg_front,
            "avg_back_hit": avg_back,
            "avg_total_hit": avg_total,
            "avg_accuracy": avg_acc,
            "best_total_hit": best["total_hit"],
            "best_accuracy": best["accuracy"],
            "best_rank": best["rank"],
            "cover_front": cover_front,
            "cover_back": cover_back,
            "cover_total": cover_front + cover_back,
            "cover_accuracy": (cover_front + cover_back) / 7.0,
        },
        "correction": correction,
    }
    _append_compare_log(result)
    return result


def _append_compare_log(result: dict, path: Path = COMPARE_LOG_PATH) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    log = []
    if path.exists():
        try:
            with path.open("r", encoding="utf-8") as f:
                log = json.load(f)
        except Exception:
            log = []
    entry = {
        "actual_period": result.get("actual_period"),
        "target_period": result.get("target_period"),
        "best_total_hit": result["summary"]["best_total_hit"],
        "best_accuracy": result["summary"]["best_accuracy"],
        "avg_accuracy": result["summary"]["avg_accuracy"],
        "cover_accuracy": result["summary"]["cover_accuracy"],
        "best_rank": result["summary"]["best_rank"],
    }
    corr = result.get("correction")
    if corr:
        entry["number_rank"] = [ALGO_NAMES[k] for k in corr["number_rank"]]
        entry["combo_rank"] = [ALGO_NAMES[k] for k in corr["combo_rank"]]
        entry["new_number_weights"] = corr["new_weights"]["number"]
        entry["new_combo_weights"] = corr["new_weights"]["combo"]
    log.append(entry)
    with path.open("w", encoding="utf-8") as f:
        json.dump(log[-100:], f, ensure_ascii=False, indent=2)


def format_compare_text(result: dict) -> str:
    s = result["summary"]
    lines = [
        f"开奖期号：{result.get('actual_period')}    预测目标期号：{result.get('target_period')}",
        f"开奖号码：前区 {fmt_nums(result['actual_front'])}  后区 {fmt_nums(result['actual_back'])}",
        "",
        f"【Top10 最佳一注】第 {s['best_rank']} 名：命中 {s['best_total_hit']}/7"
        f"（前区{result['best']['front_hit']}/5 + 后区{result['best']['back_hit']}/2）"
        f"  准确率 {s['best_accuracy']:.1%}",
        f"命中号码：前区 {fmt_nums(result['best']['front_hit_nums']) or '无'}  "
        f"后区 {fmt_nums(result['best']['back_hit_nums']) or '无'}",
        "",
        f"【Top10 平均】前区命中 {s['avg_front_hit']:.2f}/5  后区命中 {s['avg_back_hit']:.2f}/2  "
        f"合计 {s['avg_total_hit']:.2f}/7  平均准确率 {s['avg_accuracy']:.1%}",
        f"【Top10 号码池覆盖】开奖7个号中有 {s['cover_total']}/7 出现在推荐池内"
        f"（前区{s['cover_front']}/5 后区{s['cover_back']}/2）  覆盖率 {s['cover_accuracy']:.1%}",
        "",
        "各排名明细：",
    ]
    if not result.get("period_match", True):
        lines.insert(2, "注意：录入期号与上次预测目标期号不一致，仍按上次 Top10 对比。")
    for r in result["rows"]:
        lines.append(
            f"  #{r['rank']}  前区 {fmt_nums(r['front'])}  后区 {fmt_nums(r['back'])}  "
            f"→ 命中 {r['total_hit']}/7（前{r['front_hit']} 后{r['back_hit']}） "
            f"准确率 {r['accuracy']:.1%}"
        )
    corr = result.get("correction")
    if corr:
        lines.extend(["", "【算法准确率排名 → 权重修正】"])
        for row in corr["ranking_rows"]:
            sign = "+" if row["delta"] >= 0 else ""
            lines.append(
                f"  [{row['group']}] #{row['rank']} {row['name']}  "
                f"命中{row['total_hit']}/7 准确率{row['accuracy']:.1%}  "
                f"权重 {row['old_weight']:.1%} → {row['new_weight']:.1%} ({sign}{row['delta']:.1%})"
            )
        lines.append(f"累计修正次数：{corr['updates']}")
    return "\n".join(lines)


def import_excel(xlsx_path: str | Path) -> List[Draw]:
    """从 Excel 导入（列顺序：期号、日期、前区、后区）。"""
    try:
        from openpyxl import load_workbook
    except ImportError as e:
        raise RuntimeError("请先安装 openpyxl：pip install openpyxl") from e

    def parse_nums(s) -> List[int]:
        text = str(s)
        for ch in ["，", ",", " ", "-", "|", "/", "\t"]:
            text = text.replace(ch, "、")
        parts = [p for p in text.split("、") if p.strip()]
        out = []
        for p in parts:
            digits = "".join(c for c in p if c.isdigit())
            if digits:
                out.append(int(digits))
        return out

    def cell_str(v) -> str:
        if v is None:
            return ""
        if hasattr(v, "strftime"):
            return v.strftime("%Y-%m-%d")
        return str(v).strip()

    wb = load_workbook(Path(xlsx_path), read_only=True, data_only=True)
    ws = wb.active
    existing = {int(d["period"]): d for d in load_history()}
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if not row or len(row) < 4:
            continue
        period_v, date_v, front_v, back_v = row[0], row[1], row[2], row[3]
        if i == 0 and (period_v is None or not str(period_v).replace(".", "").isdigit()):
            continue  # 跳过表头
        if period_v is None or front_v is None or back_v is None:
            continue
        try:
            period = int(float(period_v))
            front = _norm_nums(parse_nums(front_v), 1, FRONT_MAX, FRONT_PICK)
            back = _norm_nums(parse_nums(back_v), 1, BACK_MAX, BACK_PICK)
            date = cell_str(date_v)[:10]
            existing[period] = {"period": period, "date": date, "front": front, "back": back}
        except Exception:
            continue
    wb.close()
    draws = sorted(existing.values(), key=lambda x: int(x["period"]))
    save_history(draws)
    return draws


if __name__ == "__main__":
    draws = load_history()
    print(f"历史 {len(draws)} 期")
    if draws:
        print("最新", draws[-1]["period"], fmt_nums(draws[-1]["front"]), "+", fmt_nums(draws[-1]["back"]))
    model = EnsemblePredictor()
    model.fit(draws)
    for i, (f, b, p, d) in enumerate(model.predict(8), 1):
        print(f"{i}. 前区 {fmt_nums(f)}  后区 {fmt_nums(b)}  {p:.2%}")
